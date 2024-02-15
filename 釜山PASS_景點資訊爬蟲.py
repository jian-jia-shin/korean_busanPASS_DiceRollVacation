from openpyxl import load_workbook
from openpyxl import Workbook
############################################
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver import Keys
from selenium.common import NoSuchElementException

import time
######################################################################################
s = Service(r"C:\chromedriver\chromedriver.exe")  # 驅動器位置(需確認chromedriver.exe放置的位置)
option = webdriver.ChromeOptions()
option.add_argument("headless")  ##執行爬蟲時不開啟瀏覽器
option.add_argument('--ignore-certificate-errors')
###遇到"你的連線不是私人連線"
###加上option.add_argument('--ignore-certificate-errors')
driver = webdriver.Chrome(service=s, options=option)
driver.implicitly_wait(4)
######################################################################################
wb = load_workbook("釜山PASS_提供之景點資料.xlsx")
wsA=wb["Group A"]
wsB=wb["Group B"]
# print(wsA.cell(row=2,column=1).value)##釜山兒童博物館
# print(wsA.cell(row=14,column=1).value==None)##True

wb_files = Workbook()##建立excel檔
wb_file_name="釜山PASS_提供之景點詳細資訊.xlsx"
wb_files.create_sheet('Group A')
wb_files.create_sheet('Group B')
wb_files.remove(wb_files['Sheet'])

wbfilesA=wb_files['Group A']##操作Group A表單
wbfilesB=wb_files['Group B']##操作Group B表單

##建立景點欄位名稱
listtop=["景點名稱","景點名稱搜尋","景點資訊","公休資訊","開放時間","停車資訊","價格","優惠","電話","地址","景點連結"]
wbfilesA.append(listtop)
wbfilesB.append(listtop)

############################################3
url_top="https://www.visitbusanpass.com/"
driver.get(url_top)
driver.implicitly_wait(5)
try:
    skip_bt=driver.find_element(By.XPATH,"/html/body/div[2]/div[1]/div/div/footer/div/button[1]")
    skip_bt.click()
except NoSuchElementException:
    print()
time.sleep(3)
find_lang_change=driver.find_element(By.ID,"__BVID__16")
find_lang_change.click()
time.sleep(3)
change_lang=driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div/header/div[1]/div/div[2]/div[1]/select/option[4]")
change_lang.click()##更換語言為中文(繁體)
time.sleep(3)
click_menu=driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div/header/div[1]/div/div[1]/button")
click_menu.click()
time.sleep(3)
to_attraction=driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div/header/div[3]/div/div/div[2]/div/ul/li[4]/div/div/header/a")
to_attraction.click()##進入景點
time.sleep(3)
##############################################
url="https://visitbusanpass.com/attractions"

def in_crawl(using_wsfile,startAt,putTo_wb_file):
    while True:
        if using_wsfile.cell(row=startAt, column=2).value == None:
            break
        else:
            driver.implicitly_wait(20)
            time.sleep(5)

            texting = driver.find_element(By.ID, "searchInput")
            texting.send_keys(using_wsfile.cell(row=startAt, column=2).value)
            time.sleep(5)
            search_ent = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div/section[1]/div/div[3]/button")
            search_ent.click()
            time.sleep(5)

            try:
                driver.implicitly_wait(7)
                enter_file = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div/div/section[2]/div[2]/div[1]/div[1]")
                enter_file.click()
                time.sleep(5)

                ##進入頁面時
                driver.implicitly_wait(20)
                time.sleep(5)
                get_info = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/section[3]/div[1]/p")
                get_dayoff = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/section[3]/div[2]/p")
                get_openrating = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div/div[1]/section[3]/div[3]/p")
                get_parking = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/section[3]/div[4]/p")
                get_price = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/section[3]/div[5]/div")
                get_benefits = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div/div[1]/section[3]/div[6]/div")
                get_phoneNum = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div/div[1]/section[3]/div[7]/div/p")
                get_address = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/section[3]/div[8]/p")
                get_website = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/section[3]/div[9]/a")

                puton = [using_wsfile.cell(row=startAt, column=1).value, using_wsfile.cell(row=startAt, column=2).value,
                         get_info.text, get_dayoff.text, get_openrating.text, get_parking.text, get_price.text,
                         get_benefits.text,get_phoneNum.text, get_address.text, get_website.text]
                putTo_wb_file.append(puton)
                wb_files.save(wb_file_name)
                driver.back()
            except NoSuchElementException:
                puton = [using_wsfile.cell(row=startAt, column=1).value, using_wsfile.cell(row=startAt, column=2).value,
                         "無相關資訊", "無相關資訊", "無相關資訊", "無相關資訊", "無相關資訊",
                         "無相關資訊","無相關資訊", "無相關資訊", "無相關資訊"]
                putTo_wb_file.append(puton)
                wb_files.save(wb_file_name)
            startAt += 1
            time.sleep(5)


in_crawl(using_wsfile=wsA,startAt=2,putTo_wb_file=wbfilesA)
in_crawl(using_wsfile=wsB,startAt=2,putTo_wb_file=wbfilesB)
driver.close()